"""
processors.py  –  Merchant Report Generator
============================================

KEY DESIGN DECISIONS:
- Address1 and Address3 are now fetched from CBS
- QR-enabled = Fonepay + Nepalpay (combined)
- Online-enabled = 0 (empty) everywhere
- Improved Local Level mapping with empty municipality handling
"""

import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CANONICAL GEOGRAPHY
# ─────────────────────────────────────────────────────────────────────────────

PROVINCES = [
    'Koshi', 'Madhesh', 'Bagmati', 'Gandaki',
    'Lumbini', 'Karnali', 'Sudurpaschim',
]

DISTRICTS = {
    'Koshi':        ['Taplejung', 'Panchthar', 'Ilam', 'Jhapa', 'Morang',
                     'Sunsari', 'Dhankuta', 'Terhathum', 'Sankhuwasabha',
                     'Bhojpur', 'Solukhumbu', 'Okhaldhunga', 'Khotang', 'Udayapur'],
    'Madhesh':      ['Sarlahi', 'Rautahat', 'Bara', 'Parsa', 'Mahottari',
                     'Dhanusha', 'Siraha', 'Saptari'],
    'Bagmati':      ['Kathmandu', 'Lalitpur', 'Bhaktapur', 'Kavrepalanchok',
                     'Sindhuli', 'Ramechhap', 'Dolakha', 'Sindhupalchok',
                     'Rasuwa', 'Nuwakot', 'Dhading', 'Makwanpur', 'Chitwan'],
    'Gandaki':      ['Kaski', 'Lamjung', 'Tanahun', 'Gorkha', 'Manang',
                     'Mustang', 'Myagdi', 'Baglung', 'Parbat', 'Syangja', 'Nawalpur'],
    'Lumbini':      ['Rupandehi', 'Kapilvastu', 'Nawalparasi West',
                     'Arghakhanchi', 'Gulmi', 'Palpa', 'Dang', 'Pyuthan',
                     'Rolpa', 'Rukum East', 'Banke', 'Bardiya'],
    'Karnali':      ['Surkhet', 'Dailekh', 'Jajarkot', 'Rukum West',
                     'Salyan', 'Dolpa', 'Humla', 'Jumla', 'Kalikot', 'Mugu'],
    'Sudurpaschim': ['Kanchanpur', 'Kailali', 'Achham', 'Doti', 'Bajhang',
                     'Bajura', 'Dadeldhura', 'Baitadi', 'Darchula'],
}

DISTRICT_TO_PROVINCE = {
    dist: prov
    for prov, dists in DISTRICTS.items()
    for dist in dists
}

PROVINCE_ALIASES = {
    'province no. 1':    'Koshi',
    'province 1':        'Koshi',
    'koshi pradesh':     'Koshi',
    'province no. 2':    'Madhesh',
    'madhedh pradesh':   'Madhesh',
    'madhes':            'Madhesh',
    'province no. 3':    'Bagmati',
    'bagmati pradesh':   'Bagmati',
    'province no. 4':    'Gandaki',
    'gandaki pradesh':   'Gandaki',
    'province no. 5':    'Lumbini',
    'lumbini pradesh':   'Lumbini',
    'province no. 6':    'Karnali',
    'karnali pradesh':   'Karnali',
    'province no. 7':    'Sudurpaschim',
    'sudurpashchim':     'Sudurpaschim',
    'far-western':       'Sudurpaschim',
}

DISTRICT_ALIASES = {
    'nawalparasi west':            'Nawalparasi West',
    'nawalparasi (west of bardaghat susta)': 'Nawalparasi West',
    'nawalparasi (east of bardaghat susta)': 'Nawalparasi East',
    'rukum (west)':                'Rukum West',
    'rukum (east)':                'Rukum East',
    'rukum east district':         'Rukum East',
    'eastern rukum':               'Rukum East',
    'western rukum':               'Rukum West',
}

LOCAL_LEVELS = [
    'Metropolitan Cities',
    'Sub-Metropolitan Cities',
    'Municipalities',
    'Rural Municipalities',
]

# ─────────────────────────────────────────────────────────────────────────────
#  NORMALISATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _normalize_province(raw: str) -> str:
    if not raw or str(raw).strip().lower() in ('', 'nan', 'none', 'null', 'n/a'):
        return ''  # ❗ DO NOT MAP

    s = str(raw).strip().lower()

    # Numeric mappings
    numeric_map = {
        '1': 'Koshi', '01': 'Koshi', '001': 'Koshi',
        '2': 'Madhesh', '02': 'Madhesh', '002': 'Madhesh',
        '3': 'Bagmati', '03': 'Bagmati', '003': 'Bagmati',
        '4': 'Gandaki', '04': 'Gandaki', '004': 'Gandaki',
        '5': 'Lumbini', '05': 'Lumbini', '005': 'Lumbini',
        '6': 'Karnali', '06': 'Karnali', '006': 'Karnali',
        '7': 'Sudurpaschim', '07': 'Sudurpaschim', '007': 'Sudurpaschim',
    }

    if s in numeric_map:
        return numeric_map[s]

    # Handle "state-1", "state 1"
    state_match = re.search(r'(state|province)[\s\-]*([1-7])', s)
    if state_match:
        return numeric_map.get(state_match.group(2), '')

    # Existing alias logic
    if s in PROVINCE_ALIASES:
        return PROVINCE_ALIASES[s]

    for prov in PROVINCES:
        if prov.lower() == s:
            return prov

    return ''  # ❗ DO NOT return raw junk

def _normalize_district(raw: str) -> str:
    if not raw or str(raw).strip().lower() in ('', 'nan', 'none', 'null', 'n/a'):
        return ''
    s = str(raw).strip()
    s_clean = re.sub(r'\s*district\s*$', '', s, flags=re.IGNORECASE).strip()
    for dist in DISTRICT_TO_PROVINCE:
        if dist.lower() == s_clean.lower():
            return dist
    lower = s_clean.lower()
    if lower in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[lower]
    for dist in sorted(DISTRICT_TO_PROVINCE, key=len, reverse=True):
        if dist.lower() in lower:
            return dist
    return s_clean

def _normalize_gender(raw: str) -> str:
    if raw is None:
        return 'Company'
    
    s = str(raw).strip().upper()

    if s in ('', 'NAN', 'NONE', 'NULL', 'N/A'):
        return 'Company'

    if s == 'M':
        return 'Male'
    if s == 'F':
        return 'Female'
    if s in ('COMPANY', 'C', 'COMP'):
        return 'Company'

    # Anything unknown → Company
    return 'Company'

def _get_local_level(municipality: str) -> str:
    """
    Improved Local Level mapping.
    Returns None for empty municipalities so they are excluded from totals.
    """
    if pd.isna(municipality) or str(municipality).strip() == '':
        return None

    m = str(municipality).strip().upper()

    # Metropolitan Cities
    if ('METRO' in m or 'MP' in m) and 'SUB' not in m:
        return 'Metropolitan Cities'
    
    # Sub-Metropolitan Cities
    if 'SUB-MP' in m or 'SUB MP' in m or 'SUBMETRO' in m:
        return 'Sub-Metropolitan Cities'
    
    # Rural Municipalities
    if 'RM' in m or 'RURAL' in m:
        return 'Rural Municipalities'
    
    # Municipalities (MC or general)
    if 'MC' in m or 'MUNICIPALITY' in m or 'MUNICIPAL' in m:
        return 'Municipalities'
    
    return 'Municipalities'


# ─────────────────────────────────────────────────────────────────────────────
#  CBS LOOKUP – Now includes Address1 and Address3
# ─────────────────────────────────────────────────────────────────────────────



def _lookup_fonepay_cbs(merchant_id: str) -> dict:
    from core.models import FonepayMerchantCBS
    empty = {
        'province': '', 
        'district': '', 
        'municipality': '', 
        'gender': '',           # Keep empty so normalize can decide
        'address1': '', 
        'address3': ''
    }
    
    if not merchant_id or str(merchant_id).strip() == '':
        return empty

    mid = str(merchant_id).strip()
    
    try:
        # Use .filter().first() and be case-insensitive if needed
        obj = FonepayMerchantCBS.objects.filter(merchant_id=mid).first()
        
        if obj:
            return {
                'province':     getattr(obj, 'province', ''),
                'district':     getattr(obj, 'district', ''),
                'municipality': getattr(obj, 'municipality', ''),
                'gender':       getattr(obj, 'gender', ''),        # ← Must come from DB
                'address1':     getattr(obj, 'address1', '') or '',
                'address3':     getattr(obj, 'address3', '') or '',
            }
    except Exception as e:
        print(f"[DEBUG] Fonepay lookup error for {mid}: {e}")
    
    return empty


def _lookup_nepalpay_cbs(merchant_code: str) -> dict:
    from core.models import NepalpayMerchantCBS
    empty = {
        'province': '', 
        'district': '', 
        'municipality': '', 
        'gender': '', 
        'address1': '', 
        'address3': '', 
        'merchant_accounts': []
    }
    
    if not merchant_code or str(merchant_code).strip() == '':
        return empty

    mc = str(merchant_code).strip()
    
    try:
        objs = list(NepalpayMerchantCBS.objects.filter(merchant_code=mc))
        if objs:
            obj = objs[0]   # Use first row for geo + gender
            accounts = [getattr(o, 'merchant_account', '') or '' for o in objs if getattr(o, 'merchant_account', '')]
            return {
                'province':     getattr(obj, 'province', ''),
                'district':     getattr(obj, 'district', ''),
                'municipality': getattr(obj, 'municipality', ''),
                'gender':       getattr(obj, 'gender', ''),      # ← Must come from DB
                'address1':     getattr(obj, 'address1', '') or '',
                'address3':     getattr(obj, 'address3', '') or '',
                'merchant_accounts': accounts,
            }
    except Exception as e:
        print(f"[DEBUG] Nepalpay lookup error for {mc}: {e}")
    
    return empty

# ─────────────────────────────────────────────────────────────────────────────
#  PROCESSING FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _is_blank(val) -> bool:
    return val is None or str(val).strip().lower() in ('', 'nan', 'none', 'null', 'n/a')


def process_fonepay(filepath: str):
    """
    Fonepay upload already carries PROVINCE/DISTRICT/MUNICIPALITY for most rows.
    CBS enrichment strategy:
      • GENDER      – always populated from CBS (never present in upload).
      • ADDRESS1/3  – always populated from CBS.
      • PROVINCE    – filled from CBS only when blank in the upload row.
      • DISTRICT    – filled from CBS only when blank in the upload row.
      • MUNICIPALITY– filled from CBS only when blank in the upload row.
    Existing geo values are normalised but not overwritten.
    """
    df = pd.read_excel(filepath, dtype={'MERCHANT_ID': str})
    original_count = len(df)

    keep_cols = [
        'MERCHANT_ID', 'MERCHANT_NAME', 'ISSUER_NAME', 'TERMINAL_DETAILS_ID',
        'PROVINCE', 'DISTRICT', 'MUNICIPALITY', 'ORIGINAL_AMOUNT', 'PAYMENT_MODULE',
    ]
    existing = [c for c in keep_cols if c in df.columns]
    out = df[existing].copy()

    for col in ('PROVINCE', 'DISTRICT', 'MUNICIPALITY'):
        if col not in out.columns:
            out[col] = ''
    out['GENDER']   = ''
    out['ADDRESS1'] = ''
    out['ADDRESS3'] = ''

    for idx, row in out.iterrows():
        mid = str(row.get('MERCHANT_ID', '')).strip()
        cbs = _lookup_fonepay_cbs(mid)

        # PROVINCE: normalise if present, else fallback to CBS
        if _is_blank(row.get('PROVINCE')):
            out.at[idx, 'PROVINCE'] = cbs['province']
        else:
            out.at[idx, 'PROVINCE'] = _normalize_province(str(row['PROVINCE']))

        # DISTRICT: normalise if present, else fallback to CBS
        if _is_blank(row.get('DISTRICT')):
            out.at[idx, 'DISTRICT'] = cbs['district']
        else:
            out.at[idx, 'DISTRICT'] = _normalize_district(str(row['DISTRICT']))

        # MUNICIPALITY: keep if present, else fallback to CBS
        if _is_blank(row.get('MUNICIPALITY')):
            out.at[idx, 'MUNICIPALITY'] = cbs['municipality']
        # else: keep the upload value unchanged

        # GENDER, ADDRESS1, ADDRESS3 – always from CBS (never in upload)
        out.at[idx, 'GENDER'] = _normalize_gender(cbs['gender'])
        out.at[idx, 'ADDRESS1'] = cbs['address1']
        out.at[idx, 'ADDRESS3'] = cbs['address3']

    return out, original_count


def process_nepalpay(filepath: str):
    """
    Nepalpay upload has no geo columns – everything comes from CBS.
    Also surfaces the CBS merchant_account(s) as a pipe-separated
    'CBS_MERCHANT_ACCOUNTS' column so operators can cross-check.
    """
    df = pd.read_excel(filepath, dtype={'Merchant Code': str})
    original_count = len(df)

    keep_cols = [
        'Merchant Code', 'Merchant Name', 'Merchant Account',
        'Amount', 'QR Type', 'Transaction Date', 'Issuer Id',
    ]
    existing = [c for c in keep_cols if c in df.columns]
    out = df[existing].copy()

    # Ensure Merchant Account column exists (may be missing in some files)
    if 'Merchant Account' not in out.columns:
        out['Merchant Account'] = ''

    out['PROVINCE']              = ''
    out['DISTRICT']              = ''
    out['MUNICIPALITY']          = ''
    out['GENDER']                = ''
    out['ADDRESS1']              = ''
    out['ADDRESS3']              = ''
    out['CBS_MERCHANT_ACCOUNTS'] = ''  # all known accounts from CBS

    for idx, row in out.iterrows():
        mc  = str(row.get('Merchant Code', '')).strip()
        cbs = _lookup_nepalpay_cbs(mc)
        out.at[idx, 'PROVINCE']              = cbs['province']
        out.at[idx, 'DISTRICT']              = cbs['district']
        out.at[idx, 'MUNICIPALITY']          = cbs['municipality']
        out.at[idx, 'GENDER'] = _normalize_gender(cbs['gender'])
        out.at[idx, 'ADDRESS1']              = cbs['address1']
        out.at[idx, 'ADDRESS3']              = cbs['address3']
        out.at[idx, 'CBS_MERCHANT_ACCOUNTS'] = ' | '.join(cbs['merchant_accounts'])

    return out, original_count


# ─────────────────────────────────────────────────────────────────────────────
#  SAVE PROCESSED EXCEL (with red highlighting)
# ─────────────────────────────────────────────────────────────────────────────

def save_processed_excel(df: pd.DataFrame, filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processed Data"

    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    alt_fill    = PatternFill("solid", fgColor="EEF4FB")
    warn_row_fill = PatternFill("solid", fgColor="FFF0D0")
    null_cell_fill = PatternFill("solid", fgColor="FF4444")
    null_cell_font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
    data_font   = Font(name="Calibri", size=9)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    GEO_COLS = {'PROVINCE', 'DISTRICT', 'MUNICIPALITY', 'GENDER'}
    headers = list(df.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 28

    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_has_missing = any(
            str(getattr(row, col.replace(' ', '_'), '') or '').strip() in ('', 'nan', 'None')
            for col in headers if col in GEO_COLS
        )
        for ci, col in enumerate(headers, 1):
            raw_val = getattr(row, col.replace(' ', '_'), None)
            display = '' if raw_val is None or str(raw_val).strip() in ('nan', 'None') else raw_val

            cell = ws.cell(row=ri, column=ci, value=display if display != '' else None)
            cell.border = border
            cell.alignment = right if isinstance(display, (int, float)) else left

            is_geo_missing = col in GEO_COLS and (display == '' or display is None)

            if is_geo_missing:
                cell.fill = null_cell_fill
                cell.font = null_cell_font
            elif row_has_missing:
                cell.fill = warn_row_fill
                cell.font = data_font
            else:
                cell.fill = alt_fill if ri % 2 == 0 else PatternFill()
                cell.font = data_font

            if isinstance(display, float):
                cell.number_format = '#,##0.00'
            elif isinstance(display, int):
                cell.number_format = '#,##0'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for ci, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(col) + 4, 12), 32)

    wb.save(filepath)


# ─────────────────────────────────────────────────────────────────────────────
#  GENERATE FINAL REPORT – Fully Updated
# ─────────────────────────────────────────────────────────────────────────────

def generate_final_report(fonepay_df: pd.DataFrame,
                          nepalpay_df: pd.DataFrame,
                          month_name: str,
                          output_path: str):
    """
    Final Report:
      • QR-enabled = Fonepay + Nepalpay combined
      • Online-enabled = 0 (empty)
      • POS = 0
      • Empty municipalities are excluded from Local Level counts
    """
    wb = openpyxl.Workbook()

    # Common styles
    title_fill   = PatternFill("solid", fgColor="EBF3FB")
    section_fill = PatternFill("solid", fgColor="1B3A6B")
    header_fill  = PatternFill("solid", fgColor="D6E4F7")
    alt_fill     = PatternFill("solid", fgColor="F5F8FD")
    total_fill   = PatternFill("solid", fgColor="E8F0FA")
    male_fill    = PatternFill("solid", fgColor="DDEEFF")
    female_fill  = PatternFill("solid", fgColor="FFE8F0")
    company_fill = PatternFill("solid", fgColor="E8FFE8")

    thin = Side(style='thin', color='BBCCE0')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right',  vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    def _font(bold=False, size=9, color="000000"):
        return Font(bold=bold, size=size, name="Calibri", color=color)

    # Prepare DataFrames
    fp = fonepay_df.copy()
    fp.columns = [c.upper() for c in fp.columns]
    fp['ORIGINAL_AMOUNT'] = pd.to_numeric(
        fp.get('ORIGINAL_AMOUNT', pd.Series(dtype=float)), errors='coerce').fillna(0)
    for col in ('PROVINCE', 'DISTRICT', 'MUNICIPALITY', 'GENDER'):
        fp[col] = fp.get(col, '').fillna('').astype(str).str.strip()
    fp['LOCAL_LEVEL'] = fp['MUNICIPALITY'].apply(_get_local_level)

    np_df = nepalpay_df.copy()
    np_df.columns = [c.upper() for c in np_df.columns]
    if 'AMOUNT' not in np_df.columns and 'ORIGINAL_AMOUNT' in np_df.columns:
        np_df['AMOUNT'] = np_df['ORIGINAL_AMOUNT']
    np_df['AMOUNT'] = pd.to_numeric(
        np_df.get('AMOUNT', pd.Series(dtype=float)), errors='coerce').fillna(0)
    for col in ('PROVINCE', 'DISTRICT', 'MUNICIPALITY', 'GENDER'):
        np_df[col] = np_df.get(col, '').fillna('').astype(str).str.strip()
    np_df['LOCAL_LEVEL'] = np_df['MUNICIPALITY'].apply(_get_local_level)

    # Helpers
    def _cell(ws, row, col, value, font=None, fill=None, num_fmt=None, align=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or _font()
        c.fill = fill or PatternFill()
        c.border = bdr
        c.alignment = align or (right if isinstance(value, (int, float)) else left)
        if num_fmt:
            c.number_format = num_fmt
        elif isinstance(value, float):
            c.number_format = '#,##0.00'
        elif isinstance(value, int):
            c.number_format = '#,##0'
        return c

    def _section_header(ws, row, start_col, label):
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 5)
        c = ws.cell(row=row, column=start_col, value=label)
        c.font = _font(bold=True, size=10, color="FFFFFF")
        c.fill = section_fill
        c.alignment = center
        c.border = bdr

    def _col_headers(ws, row, start_col, col1_label):
        for sc, lbl, span in [
            (start_col, col1_label, 2),
            (start_col + 2, 'Txn Count(Number)', 2),
            (start_col + 4, 'Txn Amount(NPR)', 2),
        ]:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=sc + span - 1)
            c = ws.cell(row=row, column=sc, value=lbl)
            c.font = _font(bold=True, color="1B3A6B")
            c.fill = header_fill
            c.alignment = center
            c.border = bdr

    def _data_trio(ws, row, start_col, label, cnt, amt, fill):
        pairs = [
            (start_col, label, fill, left),
            (start_col + 2, cnt, fill, right),
            (start_col + 4, amt, fill, right),
        ]
        for sc, val, fl, al in pairs:
            ws.merge_cells(start_row=row, start_column=sc, end_row=row, end_column=sc + 1)
            _cell(ws, row, sc, val, fill=fl, align=al)

    def _total_trio(ws, row, start_col, cnt, amt):
        _data_trio(ws, row, start_col, 'Total', cnt, amt, total_fill)
        for sc in (start_col, start_col + 2, start_col + 4):
            ws.cell(row=row, column=sc).font = _font(bold=True)

    def _agg(df_in, group_col, amount_col):
        result = {}
        for key, grp in df_in.groupby(group_col):
            result[key] = (len(grp), float(grp[amount_col].sum()))
        return result

    # =========================================================================
    # SHEET 1 – Province Wise
    # =========================================================================
    ws1 = wb.active
    ws1.title = "7.Merchant Txns_Province"
    ws1.merge_cells('A1:R1')
    c = ws1['A1']
    c.value = f"Province wise Merchant Transactions for the Month {month_name}"
    c.font = _font(bold=True, size=14, color="1B3A6B")
    c.alignment = center
    c.fill = title_fill
    ws1.row_dimensions[1].height = 28

    _section_header(ws1, 2, 1, "QR-enabled Merchants (Province-wise):")
    _section_header(ws1, 2, 7, "POS-enabled Merchants (Province-wise):")
    _section_header(ws1, 2, 13, "Online-enabled Merchants (Province-wise):")
    ws1.row_dimensions[2].height = 20

    _col_headers(ws1, 3, 1, "Province")
    _col_headers(ws1, 3, 7, "Province")
    _col_headers(ws1, 3, 13, "Province")
    ws1.row_dimensions[3].height = 20

    fp_prov = _agg(fp[fp['PROVINCE'].isin(PROVINCES)], 'PROVINCE', 'ORIGINAL_AMOUNT')
    np_prov = _agg(np_df[np_df['PROVINCE'].isin(PROVINCES)], 'PROVINCE', 'AMOUNT')

    qr_tot_c = qr_tot_a = 0
    for i, prov in enumerate(PROVINCES):
        row = 4 + i
        fill = alt_fill if i % 2 else PatternFill()
        fp_c, fp_a = fp_prov.get(prov, (0, 0.0))
        np_c, np_a = np_prov.get(prov, (0, 0.0))
        qr_count = fp_c + np_c
        qr_amount = fp_a + np_a
        qr_tot_c += qr_count
        qr_tot_a += qr_amount

        _data_trio(ws1, row, 1, prov, qr_count, qr_amount, fill)
        _data_trio(ws1, row, 7, prov, 0, 0.0, fill)
        _data_trio(ws1, row, 13, prov, 0, 0.0, fill)

    tot_row = 4 + len(PROVINCES)
    _total_trio(ws1, tot_row, 1, qr_tot_c, qr_tot_a)
    _total_trio(ws1, tot_row, 7, 0, 0.0)
    _total_trio(ws1, tot_row, 13, 0, 0.0)

    for ci in range(1, 19):
        ws1.column_dimensions[get_column_letter(ci)].width = 16

    # =========================================================================
    # SHEET 2 – District Wise
    # =========================================================================
    ws2 = wb.create_sheet("8.Merchant Txns_District")
    ws2.merge_cells('A1:R1')
    c = ws2['A1']
    c.value = f"District wise Transactions of Merchants for the Month {month_name}"
    c.font = _font(bold=True, size=14, color="1B3A6B")
    c.alignment = center
    c.fill = title_fill
    ws2.row_dimensions[1].height = 28

    _section_header(ws2, 2, 1, "QR-enabled Merchants (District-wise):")
    _section_header(ws2, 2, 7, "POS-enabled Merchants (District-wise):")
    _section_header(ws2, 2, 13, "Online-enabled Merchants (District-wise):")
    ws2.row_dimensions[2].height = 20

    _col_headers(ws2, 3, 1, "District")
    _col_headers(ws2, 3, 7, "District")
    _col_headers(ws2, 3, 13, "District")
    ws2.row_dimensions[3].height = 20

    fp_dist = _agg(fp, 'DISTRICT', 'ORIGINAL_AMOUNT')
    np_dist = _agg(np_df, 'DISTRICT', 'AMOUNT')

    all_districts = []
    for prov in PROVINCES:
        all_districts.extend(DISTRICTS[prov])

    qr_tot_c = qr_tot_a = 0
    for i, dist in enumerate(all_districts):
        row = 4 + i
        fill = alt_fill if i % 2 else PatternFill()
        fp_c, fp_a = fp_dist.get(dist, (0, 0.0))
        np_c, np_a = np_dist.get(dist, (0, 0.0))
        qr_count = fp_c + np_c
        qr_amount = fp_a + np_a
        qr_tot_c += qr_count
        qr_tot_a += qr_amount

        label = dist + ' District'
        _data_trio(ws2, row, 1, label, qr_count, qr_amount, fill)
        _data_trio(ws2, row, 7, label, 0, 0.0, fill)
        _data_trio(ws2, row, 13, label, 0, 0.0, fill)

    tot_row = 4 + len(all_districts)
    _total_trio(ws2, tot_row, 1, qr_tot_c, qr_tot_a)
    _total_trio(ws2, tot_row, 7, 0, 0.0)
    _total_trio(ws2, tot_row, 13, 0, 0.0)

    for ci in range(1, 19):
        ws2.column_dimensions[get_column_letter(ci)].width = 22

    # =========================================================================
    # SHEET 3 – Local Level Wise (Fixed)
    # =========================================================================
    ws3 = wb.create_sheet("9.Merchant Txns_Local")
    ws3.merge_cells('A1:R1')
    c = ws3['A1']
    c.value = f"Local Level Wise Merchant Transactions for the Month {month_name}"
    c.font = _font(bold=True, size=14, color="1B3A6B")
    c.alignment = center
    c.fill = title_fill
    ws3.row_dimensions[1].height = 28

    _section_header(ws3, 2, 1, "QR-enabled Merchants (Local Level-wise):")
    _section_header(ws3, 2, 7, "POS-enabled Merchants (Local Level-wise):")
    _section_header(ws3, 2, 13, "Online-enabled Merchants (Local Level-wise):")
    ws3.row_dimensions[2].height = 20

    _col_headers(ws3, 3, 1, "Local Level")
    _col_headers(ws3, 3, 7, "Local Level")
    _col_headers(ws3, 3, 13, "Local Level")
    ws3.row_dimensions[3].height = 20

    # Exclude rows where LOCAL_LEVEL is None (empty municipality)
    fp_ll = _agg(fp[fp['LOCAL_LEVEL'].notna()], 'LOCAL_LEVEL', 'ORIGINAL_AMOUNT')
    np_ll = _agg(np_df[np_df['LOCAL_LEVEL'].notna()], 'LOCAL_LEVEL', 'AMOUNT')

    qr_tot_c = qr_tot_a = 0
    for i, ll in enumerate(LOCAL_LEVELS):
        row = 4 + i
        fill = alt_fill if i % 2 else PatternFill()
        fp_c, fp_a = fp_ll.get(ll, (0, 0.0))
        np_c, np_a = np_ll.get(ll, (0, 0.0))
        qr_count = fp_c + np_c
        qr_amount = fp_a + np_a
        qr_tot_c += qr_count
        qr_tot_a += qr_amount

        _data_trio(ws3, row, 1, ll, qr_count, qr_amount, fill)
        _data_trio(ws3, row, 7, ll, 0, 0.0, fill)
        _data_trio(ws3, row, 13, ll, 0, 0.0, fill)

    tot_row = 4 + len(LOCAL_LEVELS)
    _total_trio(ws3, tot_row, 1, qr_tot_c, qr_tot_a)
    _total_trio(ws3, tot_row, 7, 0, 0.0)
    _total_trio(ws3, tot_row, 13, 0, 0.0)

    for ci in range(1, 19):
        ws3.column_dimensions[get_column_letter(ci)].width = 22

    # =========================================================================
    # SHEET 4 – Gender Wise
    # =========================================================================
    ws4 = wb.create_sheet("6.Genderwise_Merchant")
    ws4.merge_cells('A1:C1')
    c = ws4['A1']
    c.value = f"Transactions of Merchants Onboarded by Licensed Institutions-Gender Wise for the Month {month_name}"
    c.font = _font(bold=True, size=13, color="1B3A6B")
    c.alignment = center
    c.fill = title_fill
    ws4.row_dimensions[1].height = 36

    for ci, hdr in enumerate(["Gender of Proprietor(POS,QR-Enabled,Online Enabled)",
                              "Txn Count(Number)", "Txn Amount(NPR)"], 1):
        ch = ws4.cell(row=2, column=ci, value=hdr)
        ch.font = _font(bold=True, color="1B3A6B")
        ch.fill = header_fill
        ch.alignment = center
        ch.border = bdr
    ws4.row_dimensions[2].height = 22

    combined_count = {}
    combined_amt = {}
    for df_in, amt_col in [(fp, 'ORIGINAL_AMOUNT'), (np_df, 'AMOUNT')]:
        for gender, grp in df_in.groupby('GENDER'):
            g = str(gender).strip() or 'Unknown'
            combined_count[g] = combined_count.get(g, 0) + len(grp)
            combined_amt[g] = combined_amt.get(g, 0.0) + float(grp[amt_col].sum())

    gender_rows = [
        ('Male', male_fill),
        ('Female', female_fill),
        ('Others (Gender other than Male and Female)', PatternFill("solid", fgColor="F5F5F5")),
        ('Company', company_fill),
    ]

    tot_c = tot_a = 0
    for ri, (label, gfill) in enumerate(gender_rows, 3):
        key = next((k for k in combined_count if k.lower().startswith(label.split()[0].lower())), None)
        cnt = combined_count.get(key, 0) if key else 0
        amt = combined_amt.get(key, 0.0) if key else 0.0
        tot_c += cnt
        tot_a += amt

        for ci, val in enumerate([label, cnt, amt], 1):
            cc = ws4.cell(row=ri, column=ci, value=val)
            cc.font = _font()
            cc.fill = gfill
            cc.border = bdr
            cc.alignment = right if isinstance(val, (int, float)) else left
            if isinstance(val, float):
                cc.number_format = '#,##0.00'
            elif isinstance(val, int):
                cc.number_format = '#,##0'

    tot_row = 3 + len(gender_rows)
    for ci, val in enumerate(['Total', tot_c, tot_a], 1):
        cc = ws4.cell(row=tot_row, column=ci, value=val)
        cc.font = _font(bold=True)
        cc.fill = total_fill
        cc.border = bdr
        cc.alignment = right if isinstance(val, (int, float)) else left
        if isinstance(val, float):
            cc.number_format = '#,##0.00'
        elif isinstance(val, int):
            cc.number_format = '#,##0'

    ws4.column_dimensions['A'].width = 48
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 22

    wb.save(output_path)